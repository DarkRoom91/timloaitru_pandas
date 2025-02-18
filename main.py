import os
import openpyxl
import pandas as pd
from danh_sach_loai_tru import list_hdqt
from get_month import get_thang_nam, turn_to_month_year

# file excel ghi output
workbook_bieuphi_path = r"C:\Users\c291g\Desktop\test.xlsx"
# thu muc chua file 03/bhtg
file_03_folder_path = r"D:\Documents\OneDrive\DIV\Kiem Tra\ktr_2023\23_03_hong_thanh\so lieu don vi\03_bhtg\xlsx"
# ten sheet du lieu trong file 03/bhtg
bhtg03Worksheet = "Sheet1"
# header trong file 03/bhtg, 0-index
bhtg03Header = 6
# collumns will be read
bhtg03Collumns = "A, E:G,J,L"
# make list of every files 03/BHTG in folder
list_file_03 = os.listdir(path=file_03_folder_path)
list_file_03_path = []
for file_name in list_file_03:
    list_file_03_path.append(os.path.join(file_03_folder_path, file_name))
list_loai_tru = []

# --------------------def tim_loai_tru---------------------------
def tim_loai_tru(mkh_loaitru, file03):
    global list_loai_tru
    thang_nam = get_thang_nam(file03)
    for month in mkh_loaitru.month:
        if thang_nam == month:
            # noinspection PyTypeChecker
            data = pd.read_excel(file03, sheet_name=bhtg03Worksheet, header=bhtg03Header, usecols=bhtg03Collumns) #header co the thay doi
            # print(list(data.columns.values))
            frame_loai_tru = data[data["Mã khách hàng (CIF)"].isin(mkh_loaitru.mkh)]
            ngay_thang = frame_loai_tru["Ngày, tháng"].tolist()
            mkh = frame_loai_tru["Mã khách hàng (CIF)"].tolist()
            name = frame_loai_tru["Tên khách hàng"].tolist()
            name_id = frame_loai_tru["Giấy tờ cá nhân"].tolist()
            amount = frame_loai_tru["Số dư (đơn vị: đồng)"].tolist()
            phan_loai = frame_loai_tru["Phân loại tiền gửi"].tolist()
            for number in range(0, len(mkh)):
                row_list = [turn_to_month_year(date_string=ngay_thang[number]), mkh[number], name[number], name_id[number],
                            int(amount[number]), mkh_loaitru.type, phan_loai[number]]  # int(amount[i].replace(".", ""))
                print(row_list)
                list_loai_tru.append(row_list)


# loop through every file in the file 03/BHTG folder for making loai tru data
for file_path in list_file_03_path:
    for ds in list_hdqt:
        tim_loai_tru(mkh_loaitru=ds, file03=file_path)

# ghi list_loai_tru ra file excel
workbook_bieuphi = openpyxl.load_workbook(workbook_bieuphi_path)
worksheet_loaitru = workbook_bieuphi['loai_tru']
row_numt_loaitru = worksheet_loaitru.max_row
for i, j in enumerate(list_loai_tru, start=1):
    for m, n in enumerate(j, start=1):
        worksheet_loaitru.cell(i + row_numt_loaitru, m, n)
workbook_bieuphi.save(workbook_bieuphi_path)